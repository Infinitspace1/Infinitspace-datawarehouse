"""
membership_agreement_test/compute_notice_period.py

For each extracted membership-agreement JSON:
  1. Read coworker/date/renewal/termination-table info from the JSON.
  2. Query silver.nexudus_contracts for the most recent contract → floor_plan_desk_ids.
  3. Query silver.nexudus_products for those desk IDs → sum(capacity).
  4. Roll the contract forward (same-duration or monthly) until the period covers today.
  5. Look up the termination-notice table (tenure-bucket × capacity-bucket).
  6. If notice cannot be given in time, roll forward again until it can.
  7. Write everything to an Excel workbook.

Usage:
    python membership_agreement_test/compute_notice_period.py
"""
import calendar
import json
import logging
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv
load_dotenv(ROOT / ".env")

from shared.azure_clients.sql_client import get_sql_client

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
logger = logging.getLogger(__name__)

EXTRACTED_DIR = Path(__file__).resolve().parent / "extracted"
TODAY = date.today()

# ── helpers ──────────────────────────────────────────────────────────────────


_DATE_FMTS = [
    "%Y-%m-%d",    # 2025-08-01
    "%d/%m/%Y",    # 01/07/2025  or  1/3/2026
    "%d.%m.%Y",    # 14.04.2025  or  1.1.2025
    "%d/%m/%y",    # 01/03/26
    "%d.%m.%y",    # 01.03.26
]

def _clamp_day(day: int, month: int, year: int) -> int:
    """Clamp day to the valid range for the given month/year (handles 31/06, 02/31, etc.)."""
    max_day = calendar.monthrange(year, month)[1]
    return min(day, max_day)


def parse_date(val: Optional[str]) -> Optional[date]:
    if not val:
        return None
    # Strip trailing non-date text ("01.05.2025, then monthly rolling after")
    val = re.split(r"[,;]|\s+then\b", val)[0].strip()
    if not val or not any(c.isdigit() for c in val):
        return None

    # Pass 1: exact strptime match
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue

    # Pass 2: try day-clamping for invalid dates like 31/06 or 2028-02-31
    parts = re.split(r"[/.\-]", val)
    if len(parts) == 3:
        try:
            a, b, c = int(parts[0]), int(parts[1]), int(parts[2])
        except ValueError:
            return None
        # Decide if YYYY-MM-DD or DD/MM/YYYY based on first part magnitude
        if a > 31:
            y, m, d = a, b, c
        else:
            d, m, y = a, b, c
        if y < 100:
            y += 2000
        if 1 <= m <= 12 and y > 1900:
            d = _clamp_day(d, m, y)
            return date(y, m, d)

    return None


def get_renewal_category(renewal_system) -> str:
    """Normalise any renewal_system value to 'same_duration' | 'monthly' | 'personalized' | 'unknown'."""
    if isinstance(renewal_system, str):
        text = renewal_system.lower()
    elif isinstance(renewal_system, dict):
        text = (renewal_system.get("category") or "").lower()
    else:
        return "unknown"

    if any(k in text for k in [
        "same duration", "same_period", "initial commitment", "same term",
    ]):
        return "same_duration"
    if any(k in text for k in ["monthly", "month-to-month", "month to month"]):
        return "monthly"
    if text:
        return "personalized"
    return "unknown"


def renewal_system_display(renewal_system) -> str:
    if isinstance(renewal_system, str):
        return renewal_system
    if isinstance(renewal_system, dict):
        cat = renewal_system.get("category", "")
        raw = renewal_system.get("raw_text", "")
        return f"{cat} | {raw}" if raw else cat
    return str(renewal_system)


def months_between(d1: date, d2: date) -> int:
    rd = relativedelta(d2, d1)
    return rd.years * 12 + rd.months


def tenure_bucket(months: int) -> str:
    if months <= 11:
        return "1-11 months"
    if months <= 23:
        return "12-23 months"
    return "24+ months"


def capacity_bucket(cap: int) -> str:
    if cap <= 24:
        return "0-24"
    if cap <= 74:
        return "25-74"
    return "75+"


def parse_notice_months(text: str) -> int:
    m = re.match(r"(\d+)\s*month", text.lower())
    return int(m.group(1)) if m else 0


# ── renewal & notice computation ─────────────────────────────────────────────


def _renewal_delta(start: date, end: date, category: str) -> relativedelta:
    if category == "monthly":
        return relativedelta(months=1)
    return relativedelta(end, start)  # same_duration / personalized fallback


def compute_notice(
    start: date,
    end: date,
    category: str,
    cap: int,
    table: dict,
) -> tuple[date, str, int]:
    """
    Returns (real_end_date, notice_period_label, notice_months).

    Rolls the contract period forward until today fits inside,
    then checks whether notice can still be given. If not, keeps
    rolling forward until it can.
    """
    delta = _renewal_delta(start, end, category)

    current_end = end
    while current_end < TODAY:
        current_end += delta

    for _ in range(120):  # safety cap
        tenure = months_between(start, TODAY)
        row = tenure_bucket(tenure)
        col = capacity_bucket(cap)
        notice_label = (table.get(row) or {}).get(col, "N/A")
        n_months = parse_notice_months(notice_label)

        if TODAY + relativedelta(months=n_months) <= current_end:
            return current_end, notice_label, n_months

        current_end += delta

    return current_end, notice_label, n_months


# ── database queries ─────────────────────────────────────────────────────────


def most_recent_contract(sql, coworker_id: str) -> Optional[dict]:
    rows = sql.execute_query(
        "SELECT TOP 1 * FROM silver.nexudus_contracts "
        "WHERE coworker_id = ? "
        "  AND tariff_name LIKE '%Private Office%' "
        "  AND (cancellation_date IS NULL "
        "       OR CAST(start_date AS DATE) <> CAST(cancellation_date AS DATE)) "
        "ORDER BY start_date DESC",
        (int(coworker_id),),
    )
    return rows[0] if rows else None


def total_capacity(sql, floor_plan_desk_ids: str) -> int:
    cleaned = floor_plan_desk_ids.strip().strip("[]")
    ids = [x.strip() for x in cleaned.split(",") if x.strip()]
    if not ids:
        return 0
    ph = ",".join("?" * len(ids))
    rows = sql.execute_query(
        f"SELECT capacity FROM silver.nexudus_products WHERE source_id IN ({ph})",
        tuple(ids),
    )
    return sum(r.get("capacity") or 0 for r in rows)


def fetch_future_contracts(sql) -> list[dict]:
    """
    Sheet 3 data: active/future contracts with a known contract_term.
    Query taken verbatim from the brief, then aggregated per coworker_id:
      - first_start_date  : earliest start_date across all matching contracts
      - last_contract_term: contract_term of the most recent contract
      - total_tenure_months: DATEDIFF(MONTH, first_start_date, last_contract_term)
    """
    return sql.execute_query("""
        WITH future AS (
            SELECT *,
                ROW_NUMBER() OVER (PARTITION BY coworker_id ORDER BY start_date DESC) AS rn
            FROM silver.nexudus_contracts
            WHERE (
                      start_date >= GETDATE()
                   OR (start_date < GETDATE() AND ISNULL(contract_term, GETDATE()) >= GETDATE())
                  )
              AND cancellation_date IS NULL
              AND next_tariff_name LIKE '%Private Office%'
              AND ISNULL(contract_term, DATEADD(MONTH, 6, GETDATE())) >= DATEADD(MONTH, 6, GETDATE())
              AND floor_plan_desk_ids IS NOT NULL
        ),
        latest AS (
            SELECT * FROM future WHERE rn = 1
        ),
        first_po AS (
            SELECT
                coworker_id,
                MIN(start_date) AS first_start_date
            FROM silver.nexudus_contracts
            WHERE tariff_name LIKE '%Private Office%'
              AND (cancellation_date IS NULL
                   OR CAST(start_date AS DATE) <> CAST(cancellation_date AS DATE))
            GROUP BY coworker_id
        )
        SELECT
            l.coworker_id,
            l.coworker_name,
            l.coworker_company,
            l.location_name,
            f.first_start_date,
            l.contract_term AS last_contract_term,
            DATEDIFF(MONTH, f.first_start_date, l.contract_term) AS total_tenure_months
        FROM latest l
        JOIN first_po f ON l.coworker_id = f.coworker_id
        ORDER BY total_tenure_months DESC
    """)


def fetch_cancelled_coworkers(sql) -> list[dict]:
    """
    Sheet 2 data: coworkers whose most recent Private Office contract
    (excluding rows where start_date = cancellation_date) is cancelled.
    Returns first PO start_date, last cancellation_date, and total tenure.
    """
    return sql.execute_query("""
        WITH valid_po AS (
            SELECT *
            FROM silver.nexudus_contracts
            WHERE tariff_name LIKE '%Private Office%'
              AND (cancellation_date IS NULL
                   OR CAST(start_date AS DATE) <> CAST(cancellation_date AS DATE))
        ),
        ranked AS (
            SELECT *,
                ROW_NUMBER() OVER (PARTITION BY coworker_id ORDER BY start_date DESC) AS rn_last,
                ROW_NUMBER() OVER (PARTITION BY coworker_id ORDER BY start_date ASC)  AS rn_first
            FROM valid_po
        )
        SELECT
            l.coworker_id,
            l.coworker_name,
            l.coworker_company,
            l.location_name,
            f.start_date        AS first_start_date,
            l.cancellation_date AS last_cancellation_date,
            DATEDIFF(MONTH, f.start_date, l.cancellation_date) AS total_tenure_months
        FROM ranked l
        JOIN ranked f
            ON l.coworker_id = f.coworker_id AND f.rn_first = 1
        WHERE l.rn_last = 1
          AND l.cancellation_date IS NOT NULL
        ORDER BY total_tenure_months DESC
    """)


# ── Excel writer ─────────────────────────────────────────────────────────────

HEADERS_SHEET1 = [
    "Membership Agreement ID",
    "Coworker ID",
    "Coworker Name",
    "Location Name",
    "Contract Floor Plan Desk IDs",
    "Capacity",
    "Renewal System",
    "Termination Notice Table",
    "Start Date",
    "End Date",
    "Real End Date",
    "Notice Period",
    "Tenure (Months)",
    "Total Tenure (Months)",
]

KEYS_SHEET1 = [
    "membership_agreement_id",
    "coworker_id",
    "coworker_name",
    "location_name",
    "contract_floor_plan_desk_ids",
    "capacity",
    "renewal_system",
    "termination_notice_table",
    "start_date",
    "end_date",
    "real_end_date",
    "notice_period",
    "tenure_months",
    "total_tenure_months",
]

HEADERS_SHEET2 = [
    "Coworker ID",
    "Coworker Name",
    "Coworker Company",
    "Location Name",
    "First Start Date",
    "Last Cancellation Date",
    "Total Tenure (Months)",
]

KEYS_SHEET2 = [
    "coworker_id",
    "coworker_name",
    "coworker_company",
    "location_name",
    "first_start_date",
    "last_cancellation_date",
    "total_tenure_months",
]

HEADERS_SHEET3 = [
    "Coworker ID",
    "Coworker Name",
    "Coworker Company",
    "Location Name",
    "First Start Date",
    "Last Contract Term",
    "Total Tenure (Months)",
]

KEYS_SHEET3 = [
    "coworker_id",
    "coworker_name",
    "coworker_company",
    "location_name",
    "first_start_date",
    "last_contract_term",
    "total_tenure_months",
]


def _write_sheet(ws, headers: list[str], keys: list[str], rows: list[dict]):
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(*(Side(style="thin"),) * 4)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(keys, 1):
            val = row.get(key)
            if isinstance(val, datetime):
                val = val.date()
            c = ws.cell(row=ri, column=ci, value=val)
            if isinstance(val, date):
                c.number_format = "YYYY-MM-DD"
            c.border = border
            c.alignment = Alignment(
                vertical="center",
                wrap_text=(key == "termination_notice_table"),
            )

    for ci, key in enumerate(keys, 1):
        max_w = len(headers[ci - 1])
        for ri in range(2, len(rows) + 2):
            max_w = max(max_w, min(len(str(ws.cell(ri, ci).value or "")), 50))
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max_w + 3

    ws.freeze_panes = "A2"


def write_excel(
    active_rows: list[dict],
    cancelled_rows: list[dict],
    future_rows: list[dict],
    path: Path,
):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Active – Notice Periods"
    _write_sheet(ws1, HEADERS_SHEET1, KEYS_SHEET1, active_rows)

    ws2 = wb.create_sheet("Cancelled – Tenure")
    _write_sheet(ws2, HEADERS_SHEET2, KEYS_SHEET2, cancelled_rows)

    ws3 = wb.create_sheet("Future – Tenure")
    _write_sheet(ws3, HEADERS_SHEET3, KEYS_SHEET3, future_rows)

    wb.save(path)


# ── main ─────────────────────────────────────────────────────────────────────


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Compute notice periods")
    parser.add_argument(
        "--no-db", action="store_true",
        help="Skip DB queries (capacity will default to 0).",
    )
    args = parser.parse_args()

    sql = None
    if not args.no_db:
        try:
            sql = get_sql_client()
            sql.execute_query("SELECT 1")
            logger.info("Database connected successfully")
        except Exception as exc:
            logger.warning(f"Database connection failed: {exc}")
            logger.warning("Continuing without DB — capacity will default to 0")
            sql = None

    json_files = sorted(
        f for f in EXTRACTED_DIR.glob("*.json") if f.stem != "summary"
    )
    logger.info(f"Processing {len(json_files)} membership-agreement files  (today={TODAY})")

    results: list[dict] = []
    db_errors = 0

    for jf in json_files:
        data = json.loads(jf.read_text(encoding="utf-8"))
        cw_id = data.get("coworker_id", "")
        cw_name = data.get("coworker_name", "")
        file_id = data.get("file_id", "")

        p1 = data.get("page_1") or {}
        start = parse_date(p1.get("start_date"))
        end = parse_date(p1.get("end_date"))
        rs = data.get("renewal_system")
        cat = get_renewal_category(rs)
        table = data.get("termination_notice_table") or {}

        # Infer end_date for contracts with missing end_date
        if start and not end:
            opts = (p1.get("additional_commitment_term_options") or "").lower()
            raw_end = (p1.get("end_date") or "").lower()
            is_rolling = (
                cat == "monthly"
                or "rolling month" in opts
                or "rolling" in raw_end
                or opts == "rolling"
            )
            if is_rolling:
                end = start + relativedelta(months=1)
                cat = "monthly"
            elif re.search(r"(\d+)\s*month", opts):
                m = int(re.search(r"(\d+)\s*month", opts).group(1))
                end = start + relativedelta(months=m)
            else:
                end = None  # truly unknown

        # Guard: end_date before start_date is a data-extraction error
        if start and end and end < start:
            logger.warning(f"  {cw_name}: end_date ({end}) < start_date ({start}) — flagging")

        desk_ids = ""
        cap = 0
        loc_name = ""
        if sql:
            try:
                contract = most_recent_contract(sql, cw_id)
                if contract:
                    desk_ids = contract.get("floor_plan_desk_ids") or ""
                    loc_name = contract.get("location_name") or ""
                cap = total_capacity(sql, desk_ids) if desk_ids else 0
            except Exception as exc:
                logger.warning(f"  DB error for {cw_name}: {exc}")
                db_errors += 1

        real_end = notice_label = None
        n_months = tenure = total_tenure = 0
        if start and end and end > start:
            real_end, notice_label, n_months = compute_notice(
                start, end, cat, cap, table,
            )
            tenure = months_between(start, TODAY)
            total_tenure = months_between(start, real_end) if real_end else None
        elif start and end and end <= start:
            notice_label = "DATA ERROR (end <= start)"
            tenure = months_between(start, TODAY)
            total_tenure = None
        else:
            notice_label = "N/A (missing dates)"
            total_tenure = None

        results.append({
            "membership_agreement_id": file_id,
            "coworker_id": cw_id,
            "coworker_name": cw_name,
            "location_name": loc_name,
            "contract_floor_plan_desk_ids": desk_ids,
            "capacity": cap,
            "renewal_system": renewal_system_display(rs),
            "termination_notice_table": json.dumps(table) if table else "",
            "start_date": start,
            "end_date": end,
            "real_end_date": real_end,
            "notice_period": notice_label,
            "tenure_months": tenure,
            "total_tenure_months": total_tenure,
        })

        logger.info(
            f"  {cw_name:<30s}  cap={cap:<4d}  tenure={tenure:<3d}m  "
            f"notice={notice_label:<10s}  real_end={real_end}"
        )

    # Sheet 2: cancelled coworkers from DB
    cancelled: list[dict] = []
    if sql:
        try:
            cancelled = fetch_cancelled_coworkers(sql)
            logger.info(f"Fetched {len(cancelled)} cancelled coworkers from DB")
        except Exception as exc:
            logger.warning(f"Failed to fetch cancelled coworkers: {exc}")

    # Sheet 3: future/active contracts with known contract_term
    future: list[dict] = []
    if sql:
        try:
            future = fetch_future_contracts(sql)
            logger.info(f"Fetched {len(future)} future-contract coworkers from DB")
        except Exception as exc:
            logger.warning(f"Failed to fetch future contracts: {exc}")

    out = EXTRACTED_DIR / f"notice_periods_{TODAY.strftime('%Y%m%d')}.xlsx"
    write_excel(results, cancelled, future, out)
    logger.info(
        f"\nDone — Sheet 1: {len(results)} active rows, "
        f"Sheet 2: {len(cancelled)} cancelled rows, "
        f"Sheet 3: {len(future)} future rows → {out}"
    )
    if db_errors:
        logger.warning(f"  {db_errors} records had DB errors (capacity defaulted to 0)")


if __name__ == "__main__":
    main()
